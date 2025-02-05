import pandas as pd
from io import BytesIO
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


class MainData:
    def __init__(self, file_data):
        self.data = pd.read_csv(BytesIO(file_data), encoding='utf-8')
        self.data.columns = self.data.columns.str.strip()

    def get_total_shipment_quantity(self, destination_name, item_code):
        matching_rows = self.data[
            (self.data['届け先名'].str.strip() == destination_name.strip()) &
            (self.data['商品コード'].str.strip() == item_code.strip())
        ]
        return matching_rows['出荷実績検品数'].sum() if not matching_rows.empty else None


class ExcelProcessor:
    def __init__(self, file_data):
        self.file_data = file_data
        self.excel = pd.ExcelFile(BytesIO(file_data))
        self.sheet_names = self.excel.sheet_names

    def update_subsheet_shipment_quantity(self, main_data, output_file):
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            self._process_main_sheet(writer)

            for sheet in self.sheet_names[1:]:
                subsheet_data, original_headers = self._load_subsheet(sheet)
                if '商品コード' in subsheet_data.columns and '出荷数' in subsheet_data.columns:
                    self._update_shipment_quantities(subsheet_data, main_data, sheet)

                original_headers.to_excel(writer, sheet_name=sheet, index=False, header=False, startrow=0)
                subsheet_data.to_excel(writer, sheet_name=sheet, index=False, header=False, startrow=3)

        self._apply_formatting(output_file)

    def _process_main_sheet(self, writer):
        main_sheet_data = pd.read_excel(BytesIO(self.file_data), sheet_name=self.sheet_names[0], header=None)
        main_headers = main_sheet_data.iloc[0]
        main_sheet_data.columns = main_headers
        main_sheet_data = main_sheet_data[1:]
        main_sheet_data.to_excel(writer, sheet_name=self.sheet_names[0], index=False)

    def _load_subsheet(self, sheet):
        subsheet_data = pd.read_excel(BytesIO(self.file_data), sheet_name=sheet, header=None)
        original_headers = subsheet_data.iloc[:3]
        subsheet_data.columns = original_headers.iloc[2]
        subsheet_data = subsheet_data[3:]
        subsheet_data.columns = subsheet_data.columns.str.strip()
        return subsheet_data, original_headers

    def _update_shipment_quantities(self, subsheet_data, main_data, sheet_name):
        for i, row in subsheet_data.iterrows():
            item_code = str(row['商品コード']).strip()
            total_shipment = main_data.get_total_shipment_quantity(sheet_name, item_code)
            subsheet_data.at[i, '出荷数'] = total_shipment if total_shipment is not None else 0

    def _apply_formatting(self, file_path):
        original_workbook = load_workbook(BytesIO(self.file_data))  
        workbook = load_workbook(file_path)
        font_style = Font(name="MS PGothic", size=11)
        black_border = Border(
            left=Side(style="thin", color="000000"), 
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"), 
            bottom=Side(style="thin", color="000000")
        )
        dotted_border = Border(
            top=Side(style="dotted"),
            bottom=Side(style="dotted"),
            left=Side(style="dotted"),
            right=Side(style="dotted")
        )
        
        # Main Sheet Formatting
        main_sheet = workbook.active
        original_main_sheet = original_workbook.active

        max_col = main_sheet.max_column
        max_row = main_sheet.max_row
        main_sheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

        for row in range(1, original_main_sheet.max_row + 1):
            cell_value = original_main_sheet.cell(row=row, column=4).value
            main_sheet.cell(row=row, column=4, value=cell_value)

        for col_id in range(1, original_main_sheet.max_column + 1):
            col_letter = get_column_letter(col_id)
            ori_width = original_main_sheet.column_dimensions[col_letter].width
            main_sheet.column_dimensions[col_letter].width = ori_width

        # Subsheet Formatting
        for sheet_name in self.sheet_names[1:]:
            sheet = workbook[sheet_name]
            original_sheet = original_workbook[sheet_name]

            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = font_style
            sheet.freeze_panes = "A4"

            for row in sheet.iter_rows(min_row=2, values_only=False):
                for cell in row:
                    cell.border = dotted_border

            for row in sheet.iter_rows(min_row=1, max_row=3):
                for cell in row:
                    cell.border = black_border
            
            sheet.merge_cells("A1:B2")
            title_cell = sheet["A1"]
            title_cell.font = Font(name="MS PGothic", size=16, bold=True)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")

            header_fill_C1_C2 = PatternFill(start_color="ffe598", end_color="ffe598", fill_type="solid")
            header_fill_D1_D2 = PatternFill(start_color="fff2cb", end_color="fff2cb", fill_type="solid")
            header_fill_A3_D3 = PatternFill(start_color="f2f2f2", end_color="f2f2f2", fill_type="solid")

            sheet.auto_filter.ref = "C1:C" + str(sheet.max_row)

            for cell in ["C1", "C2"]:
                sheet[cell].fill = header_fill_C1_C2
            for cell in ["D1", "D2"]:
                sheet[cell].fill = header_fill_D1_D2
                sheet[cell].alignment = Alignment(horizontal="center", vertical="center")

            for cell in ["A3", "B3", "C3", "D3"]:
                sheet[cell].fill = header_fill_A3_D3
                sheet[cell].alignment = Alignment(horizontal="center", vertical="center")
            
            for col in sheet.iter_cols(min_col=3, max_col=4, min_row=4):
                for cell in col:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            for row in range(4, sheet.max_row + 1):
                sheet.row_dimensions[row].height = 14.3

            for col_idx in range(1, original_sheet.max_column + 1):
                column_letter = get_column_letter(col_idx)
                original_width = original_sheet.column_dimensions[column_letter].width
                sheet.column_dimensions[column_letter].width = original_width
        workbook.save(file_path)

st.title("Excel ファイルの更新")

main_file = st.file_uploader("会社のファイルをアップロードする (CSV)", type='csv')
excel_file = st.file_uploader("月次レポートファイルをアップロードする (Excel)", type='xlsx')

if main_file and excel_file:
    if st.button("Process Data"):
        output_file = "output.xlsx"
        main_data = MainData(main_file.read())
        excel_processor = ExcelProcessor(excel_file.read())
        excel_processor.update_subsheet_shipment_quantity(main_data, output_file)
        
        st.session_state.processed_file = output_file
        st.success("Data processed successfully!")

    if 'processed_file' in st.session_state:
        with open(st.session_state.processed_file, 'rb') as f:
            st.download_button("Download Processed Excel", f, file_name="processed_file.xlsx")
    else:
        st.warning("Please process the data first!")
